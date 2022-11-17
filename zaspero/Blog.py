from selenium import webdriver
import time
from datetime import datetime, timedelta
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
import openpyxl
import os 
import urllib.request
import re

chrome_options = webdriver.ChromeOptions()
chrome_options.add_argument('--headless')
chrome_options.add_argument('--no-sandbox')
chrome_options.add_argument("--single-process")
chrome_options.add_argument("--disable-dev-shm-usage")

service = Service(ChromeDriverManager().install())
web_driver = webdriver.Chrome(service = service, options=chrome_options)


main_url = 'http://www.avecwine.co.kr/'

wine_kor_name = [] 
wine_en_name = []

wine_type = []

wine_winery = []
wine_winery_en = []

wine_country = []
wine_region = []
wine_grape = []

wine_grade = []

wine_size = []
wine_body = []
wine_sweetness = []
wine_color = []
wine_temp = []
wine_alc = []
wine_nose = []
wine_taste = []
wine_paring = []
wine_vintage = []
wine_desc = []
winery_desc = []

wine_no = []
wine_image_list = []
wine_dec = []
num = 321

for i in range(1, 5) :

    url_list = []

    web_driver.get('https://zasperowine.com/?swoof=1&post_type=product&paged=1&paged=' + str(i))
    table = web_driver.find_element(By.XPATH, '//*[@id="woof_results_by_ajax"]/ul')
    li_list = table.find_elements(By.TAG_NAME, 'li')
    
    for a in li_list :
        a_list = a.find_element(By.TAG_NAME, 'a')
        url_list.append(a_list.get_attribute('href'))


    for url in url_list :
        web_driver.get(url)
        wine_post_id = web_driver.find_element(By.CLASS_NAME, 'zita-article').get_attribute('id').replace('post', 'product')
        wine_table = web_driver.find_element(By.XPATH, '//*[@id="' +wine_post_id +'"]')
        # 이름
        kor_name = wine_table.find_element(By.XPATH, '//*[@id="' +wine_post_id +'"]/div[2]/h1').text
        names = kor_name.split('\n')
        kor_name = names[0] + ' ' + names[1]
        en_name = wine_table.find_element(By.XPATH, '//*[@id="' +wine_post_id +'"]/div[2]/div/p').text
        wine_en_name.append(en_name)
        wine_kor_name.append(kor_name)
        
        view_desc = web_driver.find_element(By.XPATH, '//*[@id="' +wine_post_id +'"]/div[2]/div/table')
        # 규격화 된 와인 정보
        tr_list = view_desc.find_elements(By.TAG_NAME, 'tr')
        li_list = []
        for i in tr_list :
            td_list = i.find_elements(By.TAG_NAME, 'td')
            li_list.append(td_list[1].text)
        

        
        wtype = li_list[0].split('(')
        wine_type.append(wtype[0])

        wwinery = li_list[1].replace(')','').split('(')
        wine_winery.append(wwinery[0])
        wine_winery_en.append(wwinery[1])

        wcountry = li_list[2].replace(')','').split('(')
        wine_country.append(wcountry[0])
        wine_region.append(wcountry[1])

        wine_grape.append(li_list[3])

        wine_grade.append(li_list[4])
        wine_vintage.append(li_list[5])
        wine_alc.append(li_list[6])

        # 비규격 와인 정보
        body_table = web_driver.find_element(By.XPATH, '//*[@id="tab-description"]')
        desc = body_table.text.replace('\n',' ')
        desc.replace('Note','')
        desc.replace('Vivino rating','')
        desc_list = desc.split('Winery')

        wine_desc.append(desc_list[0])

        if len(desc_list) != 1 :
            winery_desc.append(desc_list[1])
        else :
            winery_desc.append('')

        # 사진
        image_url = web_driver.find_element(By.XPATH, '//*[@id="' +wine_post_id +'"]/div[1]/figure/div/a/img').get_attribute('src')

        wine_num = 'jh_' + str(num).zfill(6)
        wine_no.append(wine_num)
        image_name = '/mnt/c/Users/jake0/Desktop/Study/22sdc-1st-new-article-slack-noti/mervin/img/jh_' + str(num).zfill(6) + '_org.png'
        urllib.request.urlretrieve(image_url, image_name)

        wine_image_list.append(wine_num+'_org.png')

        num += 1

# 워크북(엑셀파일)을 새로 만듭니다.
wb = openpyxl.Workbook()

# 현재 활성화된 시트를 선택합니다.
sheet = wb.active
# A1셀에 hello world!를 입력합니다.
for index in range(len(wine_no)) : 
    i = index + 1
    sheet.cell(row = i, column = 1).value = wine_no[index]
    sheet.cell(row = i, column = 2).value = wine_en_name[index]
    sheet.cell(row = i, column = 3).value = wine_kor_name[index]
    sheet.cell(row = i, column = 4).value = wine_type[index]
    sheet.cell(row = i, column = 5).value = wine_winery[index]
    sheet.cell(row = i, column = 6).value = wine_winery_en[index]

    sheet.cell(row = i, column = 7).value = wine_country[index]
    sheet.cell(row = i, column = 8).value = wine_region[index]
    
    sheet.cell(row = i, column = 9).value = wine_grade[index]

    sheet.cell(row = i, column = 10).value = wine_grape[index]

    sheet.cell(row = i, column = 12).value = wine_vintage[index]
    sheet.cell(row = i, column = 13).value = wine_alc[index]

    sheet.cell(row = i, column = 14).value = wine_image_list[index]

    # sheet.cell(row = i, column = 9).value = wine_win[index]
    sheet.cell(row = i, column = 17).value = wine_desc[index]
    sheet.cell(row = i, column = 18).value = winery_desc[index]
    # sheet.cell(row = i, column = 12).value = wine_sweetness[index]

wb.save('zaspero.xlsx')



