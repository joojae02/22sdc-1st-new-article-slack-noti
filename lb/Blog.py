from selenium import webdriver
import time
from datetime import datetime, timedelta
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
import openpyxl
import os 
import urllib.request
import re

chrome_options = webdriver.ChromeOptions()
chrome_options.add_argument('--headless')
chrome_options.add_argument('--no-sandbox')
chrome_options.add_argument("--single-process")
chrome_options.add_argument("--disable-dev-shm-usage")

service = Service(ChromeDriverManager().install())
web_driver = webdriver.Chrome(service = service, options=chrome_options)


wine_kor_name = [] 
wine_en_name = []
wine_type = []
wine_winery = []
wine_winery_en = []

wine_country = []
wine_region = []
wine_grape = []

wine_grade = []
wine_win = []
wine_alc = []
wine_size = []
wine_body = []
wine_sweetness = []
wine_acid= []
wine_tannin = []
wine_color = []
wine_temp = []
wine_nose = []
wine_taste = []
wine_paring = []

wine_desc = []
winery_desc = []

wine_no = []
wine_image_list = []
wine_dec = []
num = 370
main_url = "http://www.lbwine.com/"
for i in range(1, 9) :
    url_list = []

    web_driver.get('http://www.lbwine.com/portfolio_list.html?tname=portfolio&page=' + str(i))
    table = web_driver.find_element(By.XPATH, '//*[@id="about"]/div[2]/div[4]')
    div_list = table.find_elements(By.CLASS_NAME, 'col-lg-3')
    
    for div in div_list :
        a_list = div.find_element(By.TAG_NAME, 'a')
        url_list.append(a_list.get_attribute('href'))


    for url in url_list :
        web_driver.get(url)
        wine_table = web_driver.find_element(By.XPATH, '//*[@id="about"]/div[2]/div[4]/div[3]/table/tbody')
        td_list = wine_table.find_elements(By.CLASS_NAME, 'board_text')

        wine_num = 'jh_' + str(num).zfill(6)
        wine_no.append(wine_num)

        wine_en_name.append(td_list[0].text)
        wine_kor_name.append(td_list[1].text)
        wine_win.append(td_list[2].text)
        tmp = td_list[3].text
        wwinery = tmp.replace(')','').split('(')
        wine_winery_en.append(wwinery[0])
        if len(wwinery) != 1 :
            wine_winery.append(wwinery[1])
        else :
            wine_winery.append('')

        wine_country.append(td_list[4].text)

        wine_type.append(td_list[5].text)
        
        wine_grape.append(td_list[6].text)

        wine_alc.append(td_list[7].text)
        wine_paring.append(td_list[8].text.replace(', ', '||'))
        num_list = []

        for i in range(9,13) :
            im_list = td_list[i].find_elements(By.TAG_NAME, 'img')
            start = 0
            for im in im_list :
                src = im.get_attribute('src')
                if 'on' in src or 'hf' in src :
                    start = start + 1
            if start == 0 :
                start = 1
            num_list.append(start)
        
        wine_sweetness.append(num_list[0])
        wine_acid.append(num_list[1])
        wine_body.append(num_list[2])
        wine_tannin.append(num_list[3])

        body_table = web_driver.find_element(By.XPATH, '//*[@id="about"]/div[2]/div[4]/div[4]/table/tbody/tr[1]/td/table/tbody/tr[1]/td/table/tbody')
        txt_desc = body_table.text.replace('\n', ' ')
        txt_desc = txt_desc.replace('DESCRIPTION', '')
        txt_d = txt_desc.split('TASTING NOTE')
        wine_desc.append(txt_d[0])
        if len(txt_d) != 1 :
            winery_desc.append(txt_d[1])
        else :
            winery_desc.append('')
        num += 1
        wine_image_list.append(wine_num+'_org.png')
'''
        image_url = web_driver.find_element(By.XPATH, '//*[@id="about"]/div[2]/div[4]/div[2]/table/tbody/tr/td/img').get_attribute('src')
    

        image_name = '/mnt/c/Users/jake0/Desktop/Study/22sdc-1st-new-article-slack-noti/img/jh_' + str(num).zfill(6) + '_org.png'
        urllib.request.urlretrieve(image_url, image_name)

        
'''
print(len(wine_num))
print(len(wine_sweetness))
print(len(wine_acid))
print(len(wine_tannin))

# 워크북(엑셀파일)을 새로 만듭니다.
wb = openpyxl.Workbook()

# 현재 활성화된 시트를 선택합니다.
sheet = wb.active
# A1셀에 hello world!를 입력합니다.
for index in range(93) : 
    i = index + 1
    # sheet.cell(row = i, column = 1).value = wine_no[index]
    sheet.cell(row = i, column = 2).value = wine_en_name[index]
    sheet.cell(row = i, column = 3).value = wine_kor_name[index]
    sheet.cell(row = i, column = 4).value = wine_type[index]
    sheet.cell(row = i, column = 5).value = wine_winery[index]
    sheet.cell(row = i, column = 6).value = wine_winery_en[index]

    sheet.cell(row = i, column = 7).value = wine_country[index]
    sheet.cell(row = i, column = 8).value = wine_win[index]

    sheet.cell(row = i, column = 9).value = wine_grape[index]
    sheet.cell(row = i, column = 11).value = wine_paring[index]

    sheet.cell(row = i, column = 10).value = wine_alc[index]

    sheet.cell(row = i, column = 12).value = wine_sweetness[index]
    sheet.cell(row = i, column = 13).value = wine_acid[index]
    sheet.cell(row = i, column = 14).value = wine_body[index]
    sheet.cell(row = i, column = 15).value = wine_tannin[index]

    sheet.cell(row = i, column = 17).value = wine_image_list[index]

    # sheet.cell(row = i, column = 9).value = wine_win[index]
    sheet.cell(row = i, column = 19).value = wine_desc[index]
    sheet.cell(row = i, column = 20).value = winery_desc[index]
    # sheet.cell(row = i, column = 12).value = wine_sweetness[index]

wb.save('lb.xlsx')

