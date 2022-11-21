import requests
from bs4 import BeautifulSoup as bs
import openpyxl
import os 
import urllib.request
import re
import time
import ssl

ssl._create_default_https_context = ssl._create_unverified_context
wine_kor_name = []
wine_en_name = []
wine_type = []
wine_winery = []
wine_winery_en = []
wine_vintage = []
wine_country = []
wine_region = []
wine_grape = []
wine_grade = []
wine_win = []
wine_alc = []
wine_size = []
wine_body = []
wine_sweetness = []
wine_acid= []
wine_tannin = []
wine_color = []
wine_temp = []
wine_nose = []
wine_taste = []
wine_paring = []

wine_desc = []
winery_desc = []

wine_no = []
wine_image_list = []
wine_dec = []
num = 463
start = 0


for i in range(1, 8) :
    url_list = []
    time.sleep(3)
    ur = "https://redwinenjy.winehero.co.kr/shop/search.php?qid=1&qbasic=1&qfrom=1000&qto=300000&page=" + str(i)
    page = requests.get(ur,verify=False)
    soup = bs(page.text, "html.parser")
    # print(soup)
    div_list = soup.find_all('div', attrs={'class': 'product_list_box'})
    # print(div_list)
    for div in div_list :
        ul = div.find('ul')
        url = ul.attrs['onclick']
        url = url.split('\'')
        url_list.append(url[1])

    for url in url_list :
        page = requests.get(url,verify=False)
        soup = bs(page.text, "html.parser")
        
        name = soup.find('div', attrs= {'class': 'product_name'}).text.replace('\n','').split('            ')
        product_container = soup.find('div', attrs = {'class': 'product_container'})
        print(num)
        kor_name = name[1]
        en_name = name[2]
        tmp = soup.find('div', attrs= {'class': 'product_all_detail'})
        ul_list = tmp.find_all('ul')
        product_te_list = soup.find_all('div', attrs={'class': 'product_te'})

        wine_kor_name.append(kor_name)
        wine_en_name.append(en_name)

        tmp = ul_list[0].text.replace('빈티지', '').replace('\n','').replace(' ','')
        wine_vintage.append(tmp)

        tmp = ul_list[1].text.replace('색', '').replace('\n','').replace(' ','')
        wine_type.append(tmp)

        tmp = ul_list[2].text.replace('원산지', '').replace('\n','').replace(' ','')
        tmp1 = tmp.split('/')
        wine_country.append(tmp1[0])
        wine_region.append(tmp)

        tmp = ul_list[3].text.replace('품종', '').replace('\n','').replace(' ','').replace(',', '||')
        wine_grape.append(tmp)

        tmp = ul_list[4].text.replace('용량', '').replace('\n','').replace(' ','')
        wine_size.append(tmp)

        if len(ul_list) == 6 :
            tmp = ul_list[5].text.replace('알코올 도수', '').replace('\n','').replace(' ','')
            wine_alc.append(tmp)
        else : 
            wine_alc.append('')

        wine_no.append(num)
        
        num = num + 1
'''
        for ex in product_te_list :
            tmp = ex.find('div', attrs={'class': 'title'}).text
            tmp1 = ex.find('div', attrs={'class': 'explanation'}).text.replace('\n',' ')
            if '테이스' in tmp :
                wine_taste[start] = tmp1
            elif '설명' in tmp :
                wine_desc[start] = tmp1
            elif '수상' in tmp :
                wine_win[start] = tmp1
            elif '생산' in tmp :
                winery_desc[start] = tmp1
'''

        
wb = openpyxl.Workbook()

# 현재 활성화된 시트를 선택합니다.
sheet = wb.active
# A1셀에 hello world!를 입력합니다.
for index in range(len(wine_no)) : 
    i = index + 1
    '''
        wine_vintage.append(vin[1])
        wine_color.append(color[1])
        wine_grape.append(grape[1])
        wine_size.append(size[1])
        wine_alc.append(alc[1])
        wine_country.append(region[1])

        wine_taste.append(explanation_list[0].text.replace('\n',' '))
        wine_desc.append(explanation_list[1].text.replace('\n',' '))
        wine_wine.append(explanation_list[2].text.replace('\n',' '))
        winery_desc.append(explanation_list[3].text.replace('\n',' '))
    '''
    # sheet.cell(row = i, column = 1).value = wine_no[index]
    sheet.cell(row = i, column = 2).value = wine_en_name[index]
    sheet.cell(row = i, column = 3).value = wine_kor_name[index]
    sheet.cell(row = i, column = 4).value = wine_type[index]
    sheet.cell(row = i, column = 5).value = wine_vintage[index]
    sheet.cell(row = i, column = 6).value = wine_size[index]
    sheet.cell(row = i, column = 7).value = wine_country[index]
    # sheet.cell(row = i, column = 8).value = wine_win[index]
    sheet.cell(row = i, column = 8).value = wine_grape[index]

    sheet.cell(row = i, column = 9).value = wine_alc[index]
    sheet.cell(row = i, column = 10).value = wine_region[index]


    #sheet.cell(row = i, column = 15).value = wine_image_list[index]

    # sheet.cell(row = i, column = 9).value = wine_win[index]
    #sheet.cell(row = i, column = 17).value = wine_taste[index]
    #sheet.cell(row = i, column = 18).value = wine_win[index]
    #sheet.cell(row = i, column = 19).value = wine_desc[index]
    #sheet.cell(row = i, column = 20).value = winery_desc[index]
    # sheet.cell(row = i, column = 12).value = wine_sweetness[index]

wb.save('redwinenjy1.xlsx')
