from selenium import webdriver
import time
from datetime import datetime, timedelta
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
import openpyxl
import os 
import urllib.request

chrome_options = webdriver.ChromeOptions()
chrome_options.add_argument('--headless')
chrome_options.add_argument('--no-sandbox')
chrome_options.add_argument("--single-process")
chrome_options.add_argument("--disable-dev-shm-usage")

service = Service(ChromeDriverManager().install())
web_driver = webdriver.Chrome(service = service, options=chrome_options)


main_url = 'http://www.avecwine.co.kr/'

wine_kor_name = []
wine_en_name = []
wine_type = []
wine_winery = []
wine_region = []
wine_grape = []
wine_grade = []
wine_win = []
wine_size = []
wine_body = []
wine_sweetness = []
wine_color = []
wine_nose = []
wine_taste = []
wine_paring = []
wine_desc = []
wine_no = []
wine_image_list = []
wine_dec = []
num = 1
for i in range(1, 13) :
    url_list = []

    web_driver.get('http://www.avecwine.co.kr/product/list.php?ca_id=&page=2&sort1=&sort2=&Projectlist2=&page=' + str(i))
    table = web_driver.find_element(By.XPATH, '//*[@id="brand_list_wrap"]/div[2]/ul')
    li_list = table.find_elements(By.TAG_NAME, 'li')
    for a in li_list :
        a_list = a.find_element(By.TAG_NAME, 'a')
        url_list.append(a_list.get_attribute('href'))


    for url in url_list :
        web_driver.get(url)
        name_table = web_driver.find_element(By.XPATH, '//*[@id="wine_view_wrap"]/div[3]/div')
        text = name_table.text
        text_list = text.split('\n')
        wine_tmp = ''
        for i in range(len(text_list)) :
            if text_list[i] == '색' or text_list[i] == ' 색':
                wine_tmp = wine_tmp + text_list[i + 1] + " "
            elif text_list[i] == '향' or text_list[i] == ' 향':
                wine_tmp = wine_tmp + text_list[i + 1] + " "
            elif text_list[i] == '맛' or text_list[i] == ' 맛':
                wine_tmp = wine_tmp + text_list[i + 1] + " "
            elif text_list[i] == '서비스 온도' or text_list[i] == ' 서비스 온도':
                wine_tmp = wine_tmp + text_list[i + 1] + " "




        wine_dec.append(name_table.text)
        '''name_table = web_driver.find_element(By.XPATH, '//*[@id="wine_view_wrap"]/div[1]/div[2]/div')
        wine_en_name.append( name_table.find_element(By.TAG_NAME, 'h3').text)
        wine_kor_name.append( name_table.find_element(By.TAG_NAME, 'span').text)

        image_url = web_driver.find_element(By.XPATH, '//*[@id="main_img1"]').get_attribute('src')
        view_desc = web_driver.find_element(By.XPATH, '//*[@id="wine_view_wrap"]/div[1]/div[2]/ul')
        li_list = view_desc.find_elements(By.TAG_NAME, 'li')
        wine_type.append(li_list[0].find_element(By.CLASS_NAME, 'w_desc').text)
        wine_winery.append(li_list[1].find_element(By.CLASS_NAME, 'w_desc').text)
        wine_region.append(li_list[2].find_element(By.CLASS_NAME, 'w_desc').text)
        wine_grape.append(li_list[3].find_element(By.CLASS_NAME, 'w_desc').text)
        wine_grade.append(li_list[4].find_element(By.CLASS_NAME, 'w_desc').text)
        wine_win.append(li_list[5].find_element(By.CLASS_NAME, 'w_desc').text)
        wine_size.append(li_list[6].find_element(By.CLASS_NAME, 'w_desc').text)

        body_table = web_driver.find_element(By.XPATH, '//*[@id="wine_view_wrap"]/div[2]/div[1]/div')
        li_list = body_table.find_elements(By.CLASS_NAME, 'on')
        wine_body.append(li_list[len(li_list) - 1].text)

        body_table = web_driver.find_element(By.XPATH, '//*[@id="wine_view_wrap"]/div[2]/div[2]/div')
        li_list = body_table.find_elements(By.CLASS_NAME, 'on')
        wine_sweetness.append(li_list[len(li_list) - 1].text)

        wine_num = 'jh_' + str(num).zfill(6)
        wine_no.append(wine_num)

        image_name = '/mnt/c/Users/jake0/Desktop/Study/22sdc-1st-new-article-slack-noti/avecwine/img/jh_' + str(num).zfill(6) + '_org.png'
        wine_image_list.append(image_name)

        num += 1'''

# 워크북(엑셀파일)을 새로 만듭니다.
wb = openpyxl.Workbook()

# 현재 활성화된 시트를 선택합니다.
sheet = wb.active
# A1셀에 hello world!를 입력합니다.
for index in range(len(wine_dec)) : 
    i = index + 1
    sheet.cell(row = i, column = 1).value = wine_dec[index]
    '''sheet.cell(row = i, column = 2).value = wine_en_name[index]
    sheet.cell(row = i, column = 3).value = wine_kor_name[index]
    sheet.cell(row = i, column = 4).value = wine_type[index]
    sheet.cell(row = i, column = 5).value = wine_winery[index]
    sheet.cell(row = i, column = 6).value = wine_region[index]
    sheet.cell(row = i, column = 7).value = wine_grape[index]
    sheet.cell(row = i, column = 8).value = wine_grade[index]
    sheet.cell(row = i, column = 9).value = wine_win[index]
    sheet.cell(row = i, column = 10).value = wine_size[index]
    sheet.cell(row = i, column = 11).value = wine_body[index]
    sheet.cell(row = i, column = 12).value = wine_sweetness[index]
    sheet.cell(row = i, column = 13).value = wine_image_list[index]'''

wb.save('wine1.xlsx')



